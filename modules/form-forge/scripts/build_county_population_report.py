"""Build county population report for counties on the coverage map."""
from __future__ import annotations

import csv
import io
import json
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
CT_LEGACY_FIPS = {
    "09001",
    "09003",
    "09005",
    "09007",
    "09009",
    "09011",
    "09013",
    "09015",
}


def load_popest(year: str) -> dict[str, dict]:
    url = (
        "https://www2.census.gov/programs-surveys/popest/datasets/"
        f"2020-{year}/counties/totals/co-est{year}-alldata.csv"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "FormForge/1.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        text = resp.read().decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    pop_cols = [c for c in reader.fieldnames if c.startswith("POPESTIMATE")]
    latest = sorted(pop_cols)[-1]
    out: dict[str, dict] = {}
    for row in reader:
        if row.get("SUMLEV") != "050":
            continue
        fips = row["STATE"].zfill(2) + row["COUNTY"].zfill(3)
        out[fips] = {
            "population": int(row[latest]),
            "county_name": row["CTYNAME"],
            "year": latest.replace("POPESTIMATE", ""),
        }
    return out


def aggregate_counties() -> list[dict]:
    counties_file = json.loads((ROOT / "data" / "city-counties.json").read_text(encoding="utf-8"))
    boot = json.loads((ROOT / "data" / "coverage-map-bootstrap.json").read_text(encoding="utf-8"))

    # Beaumont was missing county metadata in the source mapping file.
    counties_file["texas-beaumont"] = {
        "county": "Jefferson County",
        "county_fips": "48245",
        "state_code": "TX",
    }

    by_fips: dict[str, dict] = {}
    for city in boot["cities"]:
        info = counties_file.get(city["id"], {})
        fips = info.get("county_fips", "")
        if not fips:
            continue
        if fips not in by_fips:
            by_fips[fips] = {
                "state": city["state"],
                "county": info.get("county", "Unknown County"),
                "fips": fips,
                "cities_on_site": [],
            }
        by_fips[fips]["cities_on_site"].append(city["city"])

    pop_2024 = load_popest("2024")
    pop_2021 = load_popest("2021")

    results = []
    for fips, info in by_fips.items():
        use_2021 = fips in CT_LEGACY_FIPS
        pop = (pop_2021 if use_2021 else pop_2024).get(fips)
        results.append(
            {
                "state": info["state"],
                "county": info["county"],
                "fips": fips,
                "cities_on_site": len(info["cities_on_site"]),
                "city_names": ", ".join(sorted(info["cities_on_site"])),
                "population": pop["population"] if pop else None,
                "population_year": pop["year"] if pop else None,
                "population_source": (
                    "Census Vintage 2021 (legacy CT counties)"
                    if use_2021
                    else "Census Vintage 2024"
                ),
                "census_county_name": pop["county_name"] if pop else None,
            }
        )
    results.sort(key=lambda r: (r["state"], r["county"]))
    return results


def write_outputs(results: list[dict]) -> None:
    meta = {
        "description": "County populations for counties represented on Form Forge coverage map",
        "total_counties": len(results),
        "total_cities": sum(r["cities_on_site"] for r in results),
        "total_population": sum(r["population"] for r in results if r["population"]),
        "counties": results,
    }
    (ROOT / "data" / "county-populations-review.json").write_text(
        json.dumps(meta, indent=2),
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Counties"
    headers = [
        "State",
        "County",
        "FIPS",
        "Cities on Site",
        "Population",
        "Population Year",
        "Source",
        "Census County Name",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in results:
        ws.append(
            [
                row["state"],
                row["county"],
                row["fips"],
                row["cities_on_site"],
                row["population"],
                row["population_year"],
                row["population_source"],
                row["census_county_name"],
            ]
        )
    for col, width in zip("ABCDEFGH", [14, 28, 10, 12, 14, 14, 34, 28]):
        ws.column_dimensions[col].width = width
    for cell_row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in cell_row:
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"

    ws2 = wb.create_sheet("By State")
    ws2.append(["State", "Counties", "Cities on Site", "Combined County Population"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    by_state: dict[str, dict] = defaultdict(lambda: {"counties": 0, "cities": 0, "population": 0})
    for row in results:
        bucket = by_state[row["state"]]
        bucket["counties"] += 1
        bucket["cities"] += row["cities_on_site"]
        bucket["population"] += row["population"] or 0
    start = 2
    for state in sorted(by_state):
        bucket = by_state[state]
        ws2.append([state, bucket["counties"], bucket["cities"], bucket["population"]])
    end = ws2.max_row
    ws2.append(["TOTAL", f"=SUM(B{start}:B{end})", f"=SUM(C{start}:C{end})", f"=SUM(D{start}:D{end})"])
    for cell in ws2[end + 1]:
        cell.font = Font(bold=True)
    for cell_row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in cell_row:
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"

    ws3 = wb.create_sheet("Notes")
    notes = [
        ["Form Forge County Population Review"],
        [],
        ["Scope", "Counties with at least one city on the coverage map (556 cities, 10 states)."],
        ["Population source", "U.S. Census Bureau Population Estimates Program county totals."],
        ["Vintage 2024", "Used for all states except Connecticut legacy counties."],
        [
            "Vintage 2021",
            "Used for Connecticut because Census replaced legacy counties with planning regions.",
        ],
        [
            "Important",
            "Combined county population is full county population, not just covered cities.",
        ],
        ["Generated", "2026-07-05"],
    ]
    for note in notes:
        ws3.append(note)
    ws3["A1"].font = Font(bold=True, size=12)

    wb.save(ROOT / "data" / "county-populations-review.xlsx")


def main() -> None:
    results = aggregate_counties()
    write_outputs(results)
    total_pop = sum(r["population"] for r in results if r["population"])
    print(f"Counties: {len(results)}")
    print(f"Total combined county population: {total_pop:,}")
    print(f"Wrote {ROOT / 'data' / 'county-populations-review.xlsx'}")


if __name__ == "__main__":
    main()